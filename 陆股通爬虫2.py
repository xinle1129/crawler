# -*- coding: utf-8 -*-
"""
Created on Tue Nov 16 16:21:00 2021

@author: hzzxq
"""


from selenium import webdriver
from selenium.webdriver.support import expected_conditions as EC 
from selenium.webdriver.chrome.options import Options
from bs4 import BeautifulSoup
from openpyxl import Workbook 
from openpyxl import load_workbook
import os
import datetime
from tqdm import tqdm

#将A股代码转换成对应代码
def transID(a_target_id):
    if a_target_id.startswith('00'):
        return '7'+a_target_id[2:]
    elif a_target_id.startswith('300'):
        return '77'+a_target_id[3:]
    else:
        return '9'+a_target_id[3:]
    
#每日爬虫更新程序
def update_fun():
    #设置浏览器属性
    chrome_options = Options()    
    chrome_options.add_argument('--no-sandbox')#解决DevToolsActivePort文件不存在的报错
    chrome_options.add_argument('window-size=1920x3000') #指定浏览器分辨率
    chrome_options.add_argument('--disable-gpu') #谷歌文档提到需要加上这个属性来规避bug
    chrome_options.add_argument('--hide-scrollbars') #隐藏滚动条, 应对一些特殊页面
    chrome_options.add_argument('blink-settings=imagesEnabled=false') #不加载图片, 提升速度
    chrome_options.add_argument('--headless') #浏览器不提供可视化页面. linux下如果系统不支持可视化不加这条会启动失败
    chrome_options.binary_location = r"C:\Program Files\Google\Chrome\Application\chrome.exe" #手动指定使用的浏览器位置
    driver = webdriver.Chrome(options=chrome_options)
#    driver = webdriver.Chrome()
    url = 'https://www.hkexnews.hk/sdw/search/searchsdw_c.aspx'
    driver.get(url)
    #在现有的Excel中更新
    for i in range(0,1):
#    for i in range(1316,1318):
        #打开需要更新的EXCEL
        wb = load_workbook(filePath+r'\%s'%a[i])
        ws = wb.active
        a_target_id = a[i].strip('.xlsx')
        target_id = transID(a_target_id)
        ws = wb['Sheet1']
        #读取记录的最新日期等数据
        date_saved = ws.cell(row = ws.max_row,column = 3).value
        index_saved = ws.cell(row = ws.max_row,column = 1).value
        tradeCD_saved = ws.cell(row = ws.max_row,column = 2).value
        secID_saved = ws.cell(row = ws.max_row,column = 4).value
        exchangCD_saved = ws.cell(row = ws.max_row,column = 5).value
        ticker_saved = a_target_id
        ticketCode_saved = target_id
        partyName_saved = ws.cell(row = ws.max_row,column = 8).value
        
        begin = datetime.datetime.date(datetime.datetime.strptime(date_saved, "%Y-%m-%d"))
        end = datetime.date.today()
        #判断是否需要更新数据
        if (end - begin).days >= 1:
#            for j in tqdm(range(1,(end - begin).days+1)):
            for j in range(1,3):
                day = begin + datetime.timedelta(days=j)
                if day.weekday() != 6 and day.weekday() != 5 :
                        day = str(day).replace('-','/') 
                        #删除时间控件“可读”属性
                        js1="$(\"input[placeholder='年/月/日']\").removeAttr('readonly')"
                        js2="$(\"input[placeholder='年/月/日']\").attr('value','%s')"%day
                        driver.execute_script(js1)
                        driver.execute_script(js2)   
                        driver.find_element_by_xpath('//*[@id="txtStockCode"]').clear()
                        driver.find_element_by_xpath('//*[@id="txtStockName"]').clear()
                        driver.find_element_by_xpath('//*[@id="txtStockCode"]').send_keys(target_id)
                        driver.find_element_by_xpath('//*[@id="btnSearch"]').click() 
                        #判断港交所是否有当天数据需要更新
                        alert = EC.alert_is_present()(driver)
                        if  EC.alert_is_present()(driver):
                            alert.accept()
                            break
                        else :
                            #定位表格信息的位置并抓取内容
                            participant_id = driver.find_elements_by_xpath('//td[@class="col-participant-id"]/div[@class="mobile-list-body"]') 
                            participant_name = driver.find_elements_by_xpath('//td[@class="col-participant-name"]/div[@class="mobile-list-body"]') 
                            participant_address = driver.find_elements_by_xpath('//td[@class="col-address"]/div[@class="mobile-list-body"]') 
#                            participant_shareholding = driver.find_elements_by_class_name('col-shareholding.text-right')
                            participant_shareholding = driver.find_elements_by_xpath('//td[@class="col-shareholding text-right"]/div[@class="mobile-list-body"]') 
                            participant_shareholdin_percent = driver.find_elements_by_xpath('//td[@class="col-shareholding-percent text-right"]/div[@class="mobile-list-body"]') 
                            
                            #如果表格信息不为空
                            if len(participant_id) != 0 :
                                for k in range(0,len(participant_id)):  
#                                for k in range(1,2):  
                                    b = participant_shareholding[k].text
                                    holdVol = int(b.replace(',',''))
                                    index_saved = index_saved + 1
                                    endDate = str(day).replace('/','-')
                                    shcID = participant_id[k].text
                                    shcName = "中央结算系统" if participant_name[k].text=='' else participant_name[k].text
                                    address = participant_address[k].text
                                    holdPct = float(participant_shareholdin_percent[k].text.strip("%"))/100
                                    updateTime = datetime.datetime.strftime(datetime.datetime.now(),'%Y-%m-%d %H:%M:%S') 
                                    ws.append([index_saved,tradeCD_saved,endDate,secID_saved,exchangCD_saved,ticker_saved,\
                                               ticketCode_saved,partyName_saved,shcID,shcName,address,holdVol,holdPct,updateTime])
                                wb.save(filePath+r'\%s'%a[i])
                            else:
                                day = str(day).replace('/','-')
                                ws.append([day])
                                wb.save(filePath+r'\%s'%a[i])
     
            print('已经更新%s的陆股通持股信息'%a[i]) 
            print(i)
        else:
            print('%s没有需要更新的内容'%a[i])
#        wb.close()
    driver.close() 
           
            
            
if __name__ == "__main__":
    today = str(datetime.date.today()).replace('-','')
    filePath = r'.\陆股通持股明细'
    a = os.listdir(filePath)
    update_fun()
    