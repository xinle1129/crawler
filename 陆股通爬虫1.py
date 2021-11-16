# -*- coding: utf-8 -*-
"""
Created on Thu Jan 14 15:44:17 2021

@author: Chenxing Lin

"""
from selenium import webdriver
from selenium.webdriver.support import expected_conditions as EC 
from selenium.webdriver.chrome.options import Options
from bs4 import BeautifulSoup
from openpyxl import Workbook 
from openpyxl import load_workbook
import os
import datetime
from tqdm import tqdm


#每日爬虫更新程序
def update_fun():
    #设置浏览器属性
    chrome_options = Options()    
    chrome_options.add_argument('--no-sandbox')#解决DevToolsActivePort文件不存在的报错
    chrome_options.add_argument('window-size=1920x3000') #指定浏览器分辨率
    chrome_options.add_argument('--disable-gpu') #谷歌文档提到需要加上这个属性来规避bug
    chrome_options.add_argument('--hide-scrollbars') #隐藏滚动条, 应对一些特殊页面
    chrome_options.add_argument('blink-settings=imagesEnabled=false') #不加载图片, 提升速度
    chrome_options.add_argument('--headless') #浏览器不提供可视化页面. linux下如果系统不支持可视化不加这条会启动失败
    # 公司
#    chrome_options.binary_location = r"C:\Program Files\Google\Chrome\Application\chrome.exe" #手动指定使用的浏览器位置
    # 家
    chrome_options.binary_location = r"C:\Users\Yumeng\AppData\Local\Google\Chrome\Application\chrome.exe" #手动指定使用的浏览器位置
    driver = webdriver.Chrome(options=chrome_options)
#    driver = webdriver.Chrome()
    url = 'https://www.hkexnews.hk/sdw/search/searchsdw_c.aspx'
    driver.get(url)
    #在现有的Excel中更新
    for i in tqdm(range(2185,len(a))):
#    for i in range(1316,1318):
        #打开需要更新的EXCEL
        wb = load_workbook(r'H:\电脑\Python\爬虫\陆股通持股明细\运行\北上资金持股\%s'%a[i])
        ws = wb.active
        target_id = a[i].strip('.xlsx')
        ws = wb['Sheet1']
        #读取记录的最新日期
        date_saved = ws.cell(row = ws.max_row,column = 1).value
        begin = datetime.datetime.date(datetime.datetime.strptime(date_saved, "%Y-%m-%d"))
        end = datetime.date.today()
        #判断是否需要更新数据
        if (end - begin).days >= 1:
            for j in range(1,(end - begin).days+1):
                day = begin + datetime.timedelta(days=j)
                if day.weekday() != 6 and day.weekday() != 5 :
                        day = str(day).replace('-','/') 
                        #删除时间控件“可读”属性
                        js1="$(\"input[placeholder='年/月/日']\").removeAttr('readonly')"
                        js2="$(\"input[placeholder='年/月/日']\").attr('value','%s')"%day
                        driver.execute_script(js1)
                        driver.execute_script(js2)   
                        driver.find_element_by_xpath('//*[@id="txtStockCode"]').clear()
                        driver.find_element_by_xpath('//*[@id="txtStockName"]').clear()
                        driver.find_element_by_xpath('//*[@id="txtStockCode"]').send_keys(target_id)
                        driver.find_element_by_xpath('//*[@id="btnSearch"]').click() 
                        #判断港交所是否有当天数据需要更新
                        alert = EC.alert_is_present()(driver)
                        if  EC.alert_is_present()(driver):
                            alert.accept()
                            break
                        else :
                            #定位表格信息的位置并抓取内容
                            participant_id = driver.find_elements_by_class_name('col-participant-id')
                            participant_shareholding = driver.find_elements_by_class_name('col-shareholding.text-right')
                            #如果表格信息不为空
                            if len(participant_id) != 0 :
                                #查找HK机构持股总量并新建字典                   
#                                holders_value = driver.find_element_by_xpath('//*[@id="pnlResultSummary"]/div[1]/div[2]/div[2]/div[2]').text
#                                dict1 = {'于中央结算系统的持股量':int(holders_value)}
                                dict1 = {}
                                row_data = []
                                for m in range(2, ws.max_column + 1):
                                    cell_value = ws.cell(row=1, column=m).value
                                    row_data.append(cell_value)
                                for k in range(1,len(participant_id)):  
                                    b = participant_shareholding[k-1].text
                                    dict = {participant_id[k].text: int(b.replace(',',''))}
                                    dict1.update(dict)
                                if '' in dict1:
                                    dict2 = {'中央结算系统':dict1.pop('')}
                                    dict1.update(dict2)                   
                                intersection = []   
                                day = str(day).replace('/','-')
                                intersection = [day]
                                for n in row_data :
                                    intersection.append(dict1.setdefault(n,0))
                                    dict1.pop(n)
                                #判断是否有新增机构持股
                                if dict1 == {}:
                                    pass
                                else:
                                #添加新的投资者名称（列标签)
                                    mc = ws.max_column
                                    b = list(dict1.keys())
                                    for t in range(0,len(b)):
                                        ws.cell(row = 1,column = mc+t+1,value = b[t] ) 
                                        intersection.append(list(dict1.values())[t])  
                                ws.append(intersection)
                                wb.save(r'H:\电脑\Python\爬虫\陆股通持股明细\运行\北上资金持股\%s'%a[i])
                            else:
                                day = str(day).replace('/','-')
                                ws.append([day])
                                wb.save(r'H:\电脑\Python\爬虫\陆股通持股明细\运行\北上资金持股\%s'%a[i])
     
            print('已经更新%s的陆股通持股信息'%a[i]) 
            print(i)
        else:
            print('%s没有需要更新的内容'%a[i])
    driver.close()        
            
            
if __name__ == "__main__":
    today = str(datetime.date.today()).replace('-','')
    filePath = r'H:\电脑\Python\爬虫\陆股通持股明细\运行\北上资金持股'
    a = os.listdir(filePath)
    update_fun()